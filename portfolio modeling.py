from datetime import datetime, timedelta
import yfinance as yf
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from pathlib import Path

# Gather Data
tickers = []
quantidade = int(input("Digite quantas ações serão adicionadas: "))

for i in range(quantidade):
    ticker = input(f"Digite a sigla da ação {i+1} que você quer analisar: ").upper()
    tickers.append(ticker)

tnx = yf.Ticker("^TNX")

rf = (tnx.history(period="1d")["Close"].iloc[-1]) / 100

duracao = float(input("Digite o período de tempo que deseja analisar em anos (ex: 0.5 para 6 meses): ")) 

dias = int(duracao * 365)

end_date = datetime.today()
start_date = end_date - timedelta(days=dias)

start_date_str = start_date.strftime("%Y-%m-%d")
end_date_str = end_date.strftime("%Y-%m-%d")

df = yf.download(tickers, start_date, end_date, auto_adjust=True)

tickers_info = {}
returns_dict = {}

for ticker in tickers:
    tickers_info[ticker] = {}
    close_prices = df['Close'][ticker]
    monthly_prices = close_prices.resample('ME').last()
    monthly_returns = monthly_prices.pct_change().dropna()
    returns_dict[ticker] = monthly_returns
    tickers_info[ticker]["Average Monthly Returns"] = float(monthly_returns.mean())
    tickers_info[ticker]["Std Dev Monthly Returns"] = float(monthly_returns.std())
    tickers_info[ticker]["Annualized Returns Average"] = tickers_info[ticker]["Average Monthly Returns"] * 12
    tickers_info[ticker]["Annualized Returns Std Dev"] = tickers_info[ticker]["Std Dev Monthly Returns"] * 12

returns_df = pd.DataFrame(returns_dict)
cov_matrix = returns_df.cov()

# Write in excel
current_dir = Path().absolute()

filename = current_dir / "portfolio_modeling.xlsx"

if os.path.exists(filename):
    wb = load_workbook(filename)
else:
    wb = Workbook()
    wb.save(filename)
    wb = load_workbook(filename)

ws = wb.active

start_col = 5
end_col = start_col + len(tickers) - 1

for i, ticker in enumerate(tickers):
    col_letter = get_column_letter(start_col + i)
    ws[f"{col_letter}{5}"] = ticker
    ws[f"{col_letter}{6}"] = tickers_info[ticker]["Average Monthly Returns"]
    ws[f"{col_letter}{7}"] = tickers_info[ticker]["Std Dev Monthly Returns"]
    ws[f"{col_letter}{8}"] = tickers_info[ticker]["Annualized Returns Average"]
    ws[f"{col_letter}{9}"] = tickers_info[ticker]["Annualized Returns Std Dev"]


start_row = 56
end_row = start_row

tickers_list = list(cov_matrix.columns)

for i, ticker in enumerate(tickers_list):
    col_letter = get_column_letter(start_col + i)
    ws[f"{col_letter}{start_row}"] = ticker

for j, ticker in enumerate(tickers_list):
    ws[f"D{start_row + 1 + j}"] = ticker
    end_row += 1

for i, ticker_col in enumerate(tickers_list):
    for j, ticker_row in enumerate(tickers_list):
        value = cov_matrix.iloc[j, i] 
        col_letter = get_column_letter(start_col + i)
        ws[f"{col_letter}{start_row + 1 + j}"] = float(value)

ws["E14"] = rf

portfolio = f"{get_column_letter(start_col - 2)}11:{get_column_letter(end_col - 2)}11"
cov_matrix_excel = f"{get_column_letter(start_col)}57:{get_column_letter(end_col)}{end_row}"
ann_ret_avg = f"{get_column_letter(start_col)}8:{get_column_letter(end_col)}8"

ws["C15"] = f"=SUM({portfolio})"

ws["G15"] = f"=SQRT(MMULT({portfolio},MMULT({cov_matrix_excel},TRANSPOSE({portfolio}))))"
ws["H15"] = f"=MMULT({portfolio},TRANSPOSE({ann_ret_avg}))"



wb.save(filename)